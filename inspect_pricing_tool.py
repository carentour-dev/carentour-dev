#!/usr/bin/env python3
import argparse
import os
import re
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

CELL_REF_RE = re.compile(
    r"(?:(?:'([^']+)'|([A-Za-z0-9_]+))!)?\$?([A-Z]{1,3})\$?\d+",
    re.I,
)
COL_RANGE_RE = re.compile(
    r"(?:(?:'([^']+)'|([A-Za-z0-9_]+))!)?\$?([A-Z]{1,3})\$?:\$?([A-Z]{1,3})\$?",
    re.I,
)


def is_empty(value: object) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def build_header_map(worksheet, header_row: int) -> Dict[str, str]:
    header_map: Dict[str, str] = {}
    for col_idx in range(1, worksheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        raw_value = worksheet.cell(row=header_row, column=col_idx).value
        if is_empty(raw_value):
            header_map[col_letter] = f"<Column {col_letter}>"
        elif isinstance(raw_value, str):
            cleaned = raw_value.strip()
            if cleaned.startswith("="):
                header_map[col_letter] = f"<Column {col_letter}>"
            else:
                header_map[col_letter] = cleaned
        else:
            header_map[col_letter] = f"<Column {col_letter}>"
    return header_map


def extract_column_refs(formula: str, current_sheet: str) -> Set[Tuple[str, str]]:
    if not formula or not isinstance(formula, str) or not formula.startswith("="):
        return set()

    refs: Set[Tuple[str, str]] = set()
    for match in CELL_REF_RE.finditer(formula):
        sheet = (match.group(1) or match.group(2) or current_sheet).strip()
        column = match.group(3).upper()
        refs.add((sheet, column))
    for match in COL_RANGE_RE.finditer(formula):
        sheet = (match.group(1) or match.group(2) or current_sheet).strip()
        start_col = match.group(3).upper()
        end_col = match.group(4).upper()
        refs.add((sheet, start_col))
        refs.add((sheet, end_col))
    return refs


def sorted_sheet_columns(refs: Set[Tuple[str, str]]) -> List[Tuple[str, str]]:
    return sorted(refs, key=lambda item: (item[0].lower(), column_index_from_string(item[1])))


def collect_formula_cells(worksheet) -> List[Tuple[int, int, str]]:
    formulas: List[Tuple[int, int, str]] = []
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=row, column=col).value
            if is_formula(value):
                formulas.append((row, col, value))
    return formulas


def build_dependency_graph(
    worksheet,
) -> Tuple[Dict[str, Set[Tuple[str, str]]], Dict[str, int], bool, int]:
    formula_cells = collect_formula_cells(worksheet)
    dependencies_by_col: Dict[str, Set[Tuple[str, str]]] = {}
    formula_counts: Dict[str, int] = {}
    sheet_refs_found = False

    for _, col_idx, formula in formula_cells:
        col_letter = get_column_letter(col_idx)
        formula_counts[col_letter] = formula_counts.get(col_letter, 0) + 1
        dependencies_by_col.setdefault(col_letter, set())
        refs = extract_column_refs(formula, worksheet.title)
        if any(ref_sheet != worksheet.title for ref_sheet, _ in refs):
            sheet_refs_found = True
        if refs:
            dependencies_by_col[col_letter].update(refs)

    return dependencies_by_col, formula_counts, sheet_refs_found, len(formula_cells)


def find_first_formula_row(worksheet) -> Optional[int]:
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            if is_formula(worksheet.cell(row=row, column=col).value):
                return row
    return None


def find_header_row(worksheet, search_rows: int) -> int:
    max_row = min(worksheet.max_row, search_rows)
    best_row = 1
    best_count = -1
    best_has_formula = True

    for row in range(1, max_row + 1):
        values = [worksheet.cell(row=row, column=col).value for col in range(1, worksheet.max_column + 1)]
        non_empty_count = sum(1 for value in values if not is_empty(value))
        if non_empty_count == 0:
            continue
        row_has_formula = any(is_formula(value) for value in values)
        if non_empty_count > best_count or (
            non_empty_count == best_count and best_has_formula and not row_has_formula
        ):
            best_row = row
            best_count = non_empty_count
            best_has_formula = row_has_formula

    return best_row if best_count >= 0 else 1


def format_dependency(
    ref_sheet: str,
    ref_col: str,
    current_sheet: str,
    header_map_by_sheet: Dict[str, Dict[str, str]],
) -> str:
    header_map = header_map_by_sheet.get(ref_sheet, {})
    header_label = header_map.get(ref_col, f"<Column {ref_col}>")
    if ref_sheet == current_sheet:
        return f"{header_label} ({ref_col})"
    return f"{ref_sheet}!{ref_col} ({header_label})"


def analyze_sheet(
    worksheet,
    header_row: int,
    data_row: int,
    header_map_by_sheet: Dict[str, Dict[str, str]],
    header_source: str,
    data_source: str,
    scan_all_formulas: bool,
) -> None:
    max_col = worksheet.max_column
    header_map = header_map_by_sheet[worksheet.title]

    header_suffix = " (auto)" if header_source == "auto" else ""
    if data_source == "auto":
        data_suffix = " (auto)"
    elif data_source == "auto-fallback":
        data_suffix = " (auto fallback; no formulas found)"
    else:
        data_suffix = ""

    print(f"Sheet: {worksheet.title}")
    print(f"Header row: {header_row}{header_suffix}")
    print(f"Data row: {data_row}{data_suffix}")

    print(f"\nHeaders (row {header_row}):")
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        header_value = worksheet.cell(row=header_row, column=col_idx).value
        display = "<empty>" if is_empty(header_value) else header_value
        print(f"- {col_letter}: {display}")

    print(f"\nRaw formulas (row {data_row}):")
    formulas_by_col: Dict[str, str] = {}
    sheet_refs_found = False
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        cell = worksheet.cell(row=data_row, column=col_idx)
        formula = cell.value if is_formula(cell.value) else None
        if formula and "!" in formula:
            sheet_refs_found = True
        if formula:
            formulas_by_col[col_letter] = formula
            print(f"- {col_letter} ({header_map[col_letter]}): {formula}")
        else:
            print(f"- {col_letter} ({header_map[col_letter]}): <no formula>")

    print(f"\nDependencies (based on row {data_row} formulas):")
    if not formulas_by_col:
        print("- No formulas found in the first data row.")
    else:
        for col_letter in sorted(formulas_by_col.keys(), key=lambda c: column_index_from_string(c)):
            formula = formulas_by_col[col_letter]
            refs = extract_column_refs(formula, worksheet.title)
            if not refs:
                print(f"- {col_letter} ({header_map[col_letter]}): <no column refs detected>")
                continue
            ref_list = sorted_sheet_columns(refs)
            ref_labels = ", ".join(
                format_dependency(ref_sheet, ref_col, worksheet.title, header_map_by_sheet)
                for ref_sheet, ref_col in ref_list
            )
            print(f"- {col_letter} ({header_map[col_letter]}) depends on: {ref_labels}")

    if sheet_refs_found:
        if not scan_all_formulas:
            print("\nNote: cross-sheet dependencies are listed as Sheet!Column.")

    if scan_all_formulas:
        print("\nDependency graph (all formulas in sheet):")
        dependencies_by_col, formula_counts, sheet_refs_found_all, total_formulas = build_dependency_graph(
            worksheet
        )
        if total_formulas == 0:
            print("- No formulas found in sheet.")
        else:
            print(f"- Formula cells found: {total_formulas}")
            for col_letter in sorted(formula_counts.keys(), key=column_index_from_string):
                header_label = header_map.get(col_letter, f"<Column {col_letter}>")
                count = formula_counts[col_letter]
                refs = dependencies_by_col.get(col_letter, set())
                if not refs:
                    print(
                        f"- {col_letter} ({header_label}) [formulas: {count}]: "
                        "<no column refs detected>"
                    )
                    continue
                ref_list = sorted_sheet_columns(refs)
                ref_labels = ", ".join(
                    format_dependency(ref_sheet, ref_col, worksheet.title, header_map_by_sheet)
                    for ref_sheet, ref_col in ref_list
                )
                print(
                    f"- {col_letter} ({header_label}) [formulas: {count}] depends on: {ref_labels}"
                )
            if sheet_refs_found_all:
                print("\nNote: cross-sheet dependencies are listed as Sheet!Column.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect headers and formulas in the CareNTour pricing tool.")
    parser.add_argument(
        "--file",
        default="CareNTour_Pricing_Tool.xlsx",
        help="Path to the Excel file to inspect.",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Sheet name to inspect (defaults to the active sheet).",
    )
    parser.add_argument("--header-row", type=int, default=1, help="Header row number.")
    parser.add_argument("--data-row", type=int, default=2, help="First data row number.")
    parser.add_argument(
        "--auto-header-row",
        action="store_true",
        help="Auto-detect the header row based on filled cells.",
    )
    parser.add_argument(
        "--auto-data-row",
        action="store_true",
        help="Auto-detect the first row that contains formulas.",
    )
    parser.add_argument(
        "--header-search-rows",
        type=int,
        default=10,
        help="Rows to scan when auto-detecting header row.",
    )
    parser.add_argument(
        "--all-sheets",
        action="store_true",
        help="Analyze all sheets in the workbook.",
    )
    parser.add_argument(
        "--scan-all-formulas",
        action="store_true",
        help="Build a dependency graph using all formula cells in each sheet.",
    )
    args = parser.parse_args()

    file_path = args.file
    if not os.path.exists(file_path):
        raise SystemExit(f"File not found: {file_path}")

    workbook = load_workbook(file_path, data_only=False)
    worksheets = workbook.worksheets if args.all_sheets else [workbook[args.sheet] if args.sheet else workbook.active]

    auto_data_row = args.auto_data_row or args.all_sheets
    header_rows: Dict[str, int] = {}
    header_sources: Dict[str, str] = {}
    data_rows: Dict[str, int] = {}
    data_sources: Dict[str, str] = {}

    for ws in worksheets:
        if args.auto_header_row:
            header_rows[ws.title] = find_header_row(ws, args.header_search_rows)
            header_sources[ws.title] = "auto"
        else:
            header_rows[ws.title] = args.header_row
            header_sources[ws.title] = "manual"

    header_map_by_sheet = {
        ws.title: build_header_map(ws, header_rows[ws.title]) for ws in worksheets
    }

    for ws in worksheets:
        if auto_data_row:
            formula_row = find_first_formula_row(ws)
            if formula_row is None:
                data_rows[ws.title] = args.data_row
                data_sources[ws.title] = "auto-fallback"
            else:
                data_rows[ws.title] = formula_row
                data_sources[ws.title] = "auto"
        else:
            data_rows[ws.title] = args.data_row
            data_sources[ws.title] = "manual"

    for idx, ws in enumerate(worksheets):
        analyze_sheet(
            ws,
            header_rows[ws.title],
            data_rows[ws.title],
            header_map_by_sheet,
            header_sources[ws.title],
            data_sources[ws.title],
            args.scan_all_formulas,
        )
        if idx < len(worksheets) - 1:
            print("\n" + "-" * 60 + "\n")


if __name__ == "__main__":
    main()
